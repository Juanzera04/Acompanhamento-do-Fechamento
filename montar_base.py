import pandas as pd
from pathlib import Path
from openpyxl import load_workbook
from openpyxl.worksheet.table import Table, TableStyleInfo

# ==========================================
# CAMINHOS
# ==========================================
PASTA_RADAR = Path("Data/Relatorio_base_CTB")
PASTA_CHECKLIST = Path("Data/Checklist_CTB")
PASTA_RETORNO = Path("Data/Retorno_Checklist")
PASTA_COORDENADORES = Path("Data/Coordenadores_CTB")

# ==========================================
# FUNÇÃO PARA PEGAR ARQUIVO
# ==========================================
def obter_arquivo(pasta):
    arquivos = list(pasta.glob("*.xlsx"))
    if not arquivos:
        raise FileNotFoundError(f"Nenhum arquivo encontrado em {pasta}")
    return arquivos[0]

# ==========================================
# LEITURA
# ==========================================
arquivo_radar = obter_arquivo(PASTA_RADAR)
arquivo_checklist = obter_arquivo(PASTA_CHECKLIST)
arquivo_retorno = obter_arquivo(PASTA_RETORNO)
arquivo_coordenador = obter_arquivo(PASTA_COORDENADORES)
df_radar = pd.read_excel(arquivo_radar)
df_checklist = pd.read_excel(arquivo_checklist)
df_retorno = pd.read_excel(arquivo_retorno, sheet_name="Pendencias")
df_coordenador = pd.read_excel(arquivo_coordenador)

# ==========================================
# LIMPEZA RADAR
# ==========================================
colunas_remover = [
    "DataSimulacao",
    "Data",
    "DataFechamento",
    "ResponsavelManipulacao",
    "FechamentoBranco"
]

df_radar = df_radar.drop(columns=[c for c in colunas_remover if c in df_radar.columns])

# ==========================================
# AJUSTAR TRIBUTAÇÃO
# ==========================================
mapa_tributacao = {
    "Federal - Lucro Presumido": "Lucro Presumido",
    "Federal - SN": "Simples Nacional",
    "Federal - Imune": "Simples Nacional",
    "Federal - L Real -Trimestral": "Lucro Real",
    "Federal - Lucro Real - Anual": "Lucro Real",
    "Federal - L.Real - Mensal": "Lucro Real",
    "Federal - MEI": "Simples Nacional"
}

df_radar["Tributacao"] = df_radar["Tributacao"].replace(mapa_tributacao)

# ==========================================
# ===============================
# 🔹 1. STATUS DOCUMENTAÇÃO
# ===============================
# Agrupa por cliente (caso tenha mais de uma linha)
status_map = df_retorno.groupby("CodCliente")["Status"].apply(list).to_dict()

def definir_status(lista_status):
    if not lista_status:
        return ""
    # Se tiver qualquer coisa diferente de "baixado" → Pendente
    for s in lista_status:
        if str(s).strip().lower() != "baixado":
            return "Pendente"
    return "Recebida"

df_radar["StatusDocumentacao"] = df_radar["IdCliente"].map(
    lambda x: definir_status(status_map.get(x, []))
)

# ==========================================
# 🔹 2. DATA DOCUMENTAÇÃO
# ===============================
data_map = df_retorno.groupby("CodCliente")["DataBaixa"].max().to_dict()

df_radar["DataDocumentacao"] = df_radar["IdCliente"].map(data_map)

# ✅ Apenas UMA conversão, com dayfirst=True para garantir leitura correta
df_radar["DataDocumentacao"] = pd.to_datetime(
    df_radar["DataDocumentacao"], errors="coerce", dayfirst=True
).dt.strftime("%d/%m/%Y %H:%M:%S")

# ==========================================
# 🔹 3. DOCUMENTAÇÃO PENDENTE
# ===============================
doc_map = (
    df_checklist
    .groupby("IdCliente")["Tipo"]
    .apply(lambda x: " | ".join(sorted(set(map(str, x)))))
    .to_dict()
)

df_radar["DocumentacaoPendente"] = df_radar["IdCliente"].map(doc_map)

# ==========================================
# 🔹 4. COORDENAÇÃO
# ==========================================
# Cria mapa:
# Nome do colaborador -> Coordenador

coord_map = (
    df_coordenador
    .set_index("Nome de Exibição")["Coordenador"]
    .to_dict()
)

df_radar["Coordenação"] = df_radar["EquipeAtendimento"].map(coord_map)

# ==========================================
# 🔹 5. AJUSTAR SEGMENTO
# ==========================================
# Cria mapa:
# Nome do colaborador -> Departamento

departamento_map = (
    df_coordenador
    .set_index("Nome de Exibição")["Departamento"]
    .to_dict()
)

# Mapeamento departamento -> Segmento
mapa_segmento = {
    "CTB - CONTÁBIL HOLDING": "Holding",
    "CTB - CONTÁBIL INDUSTRIA": "Industria",
    "CTB - CONTÁBIL VAREJO": "Varejo",
    "SANTOS - CTB": "Varejo",
    "RJ - CTB": "Varejo"
}

def ajustar_segmento(row):
    colaborador = row["EquipeAtendimento"]

    departamento = departamento_map.get(colaborador)

    # Se encontrou departamento no mapa
    if departamento in mapa_segmento:
        return mapa_segmento[departamento]

    # Se não encontrou, mantém valor atual
    return row["Segmento"]

df_radar["Segmento"] = df_radar.apply(ajustar_segmento, axis=1)

# ==========================================
# SALVAR COM PANDAS
# ==========================================
saida = PASTA_RADAR / "Radar.xlsx"
df_radar.to_excel(saida, index=False)

# ==========================================
# FORMATAR COMO TABELA
# ==========================================
wb = load_workbook(saida)
ws = wb.active

# Define o range da tabela
max_row = ws.max_row
max_col = ws.max_column

from openpyxl.utils import get_column_letter
end_col_letter = get_column_letter(max_col)

tabela_ref = f"A1:{end_col_letter}{max_row}"

# Cria a tabela
tabela = Table(displayName="TabelaRadar", ref=tabela_ref)

# Estilo da tabela
estilo = TableStyleInfo(
    name="TableStyleMedium9",
    showFirstColumn=False,
    showLastColumn=False,
    showRowStripes=True,
    showColumnStripes=False
)

tabela.tableStyleInfo = estilo
ws.add_table(tabela)

# Salva novamente
wb.save(saida)

print(f"✅ Arquivo final formatado como tabela: {saida}")